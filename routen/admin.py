"""
Admin-Dashboard-Modul für Seelauf-App
Verwaltung von Schülern, Läufen und Zeitmessungen mit Authentifizierung
"""
from functools import wraps
import os
import hmac
import unicodedata

from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify
from openpyxl import load_workbook

import db

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')
ADMIN_CODE = os.getenv('ADMIN_CODE', '123')


def login_required(f):
    """
    Decorator: Erzwingt Admin-Login.
    Umleitet auf Login-Seite, wenn Benutzer nicht authentifiziert ist.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            flash('Bitte melden Sie sich zuerst an.', 'warning')
            return redirect(url_for('admin.login'))
        return f(*args, **kwargs)
    return decorated_function


# ============================================================================
# AUTHENTIFIZIERUNG
# ============================================================================

@admin_bp.route('/login', methods=['GET', 'POST'])
def login():
    """
    Admin-Login-Seite.
    Prüft den Zugangscode mit HMAC-Vergleich.
    """
    if request.method == 'POST':
        code = request.form.get('code', '')
        # Sichere Vergleich mit hmac.compare_digest (gegen Timing-Attacken)
        if hmac.compare_digest(code, ADMIN_CODE):
            session['admin_logged_in'] = True
            flash('Erfolgreich angemeldet.', 'success')
            return redirect(url_for('admin.dashboard'))
        flash('Falscher Zugangscode.', 'danger')
    return render_template('admin_login.html')


@admin_bp.route('/logout')
def logout():
    """Admin-Logout: Sitzung beenden und zur Login-Seite."""
    session.pop('admin_logged_in', None)
    flash('Sie wurden abgemeldet.', 'info')
    return redirect(url_for('admin.login'))


@admin_bp.route('/')
@admin_bp.route('/dashboard')
@login_required
def dashboard():
    """Admin-Dashboard: Übersicht für Administrator."""
    return render_template('admin_dashboard.html')



# ============================================================================
# SCHÜLER-VERWALTUNG
# ============================================================================

@admin_bp.route('/students')
@login_required
def list_students():
    """Schüler-Liste: Alle Schüler anzeigen."""
    class_group = request.args.get('class_group', '').strip()
    nummer = request.args.get('nummer', '').strip()
    students = db.get_students(
        include_ill=True,
        class_group=class_group or None,
        nummer=nummer or None
    )
    return render_template(
        'admin_students.html',
        students=students,
        classes=db.get_classes(),
        selected_class=class_group,
        nummer_filter=nummer
    )


@admin_bp.route('/classes', methods=['GET', 'POST'])
@login_required
def list_classes():
    """Klassen und zugehörige Lehrkräfte verwalten."""
    if request.method == 'POST':
        old_class_group = request.form.get('old_class_group', '').strip()
        class_group = request.form.get('class_group', '').strip()
        teacher = request.form.get('teacher', '').strip()
        if not old_class_group or not class_group:
            flash('Der Klassenname darf nicht leer sein.', 'danger')
        else:
            try:
                db.update_class(old_class_group, class_group, teacher)
                flash('Klasse erfolgreich aktualisiert.', 'success')
            except Exception as error:
                flash(f'Klasse konnte nicht aktualisiert werden: {error}', 'danger')
        return redirect(url_for('admin.list_classes'))

    return render_template('admin_classes.html', classes=db.get_classes())


@admin_bp.route('/add_student', methods=['GET', 'POST'])
@login_required
def add_student():
    """
    Neuen Schüler hinzufügen.
    Validiert erforderliche Felder.
    """
    if request.method == 'POST':
        class_group = request.form.get('class_group', '').strip()
        nummer = request.form.get('nummer', '').strip()
        ill = bool(request.form.get('ill'))
        
        # Validierung: Alle Felder erforderlich
        if not all([class_group, nummer]):
            flash('Klasse und Nummer müssen angegeben werden.', 'danger')
            return redirect(url_for('admin.add_student'))
        
        try:
            db.add_student(class_group, nummer, ill)
            flash('Schüler erfolgreich hinzugefügt.', 'success')
            return redirect(url_for('admin.list_students'))
        except Exception as e:
            flash(f'Fehler beim Hinzufügen: {e}', 'danger')
    
    return render_template('admin_add_student.html')


def _normalisiere_spaltenname(value):
    text = unicodedata.normalize('NFKD', str(value or ''))
    text = ''.join(char for char in text if not unicodedata.combining(char))
    return ''.join(char for char in text.lower() if char.isalnum())


def _finde_excel_spalten(workbook):
    """Findet Klasse- und Nummernspalten auch bei Vorspannzeilen."""
    class_headers = {'klasse', 'klassegruppe', 'gruppe', 'class', 'classgroup'}
    number_headers = {
        'nummer', 'startnummer', 'startnr', 'number', 'laufnummer',
        'schuelernummer', 'teilnehmernummer', 'zahl', 'nr', 'anzahl',
        'anzahlschueler', 'schueleranzahl'
    }
    for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(values_only=True))
        for header_index, candidate in enumerate(values[:20]):
            normalized = [_normalisiere_spaltenname(value) for value in candidate]
            class_index = next(
                (index for index, value in enumerate(normalized) if value in class_headers),
                None
            )
            number_index = next(
                (index for index, value in enumerate(normalized)
                 if value in number_headers or value.startswith('anzahl')),
                None
            )
            if class_index is not None and number_index is not None:
                return sheet, values, header_index, class_index, number_index
    return None


def _ist_excel_nummer(value):
    if value is None or value == '':
        return False
    if isinstance(value, (int, float)):
        return True
    return str(value).strip().replace('.', '', 1).isdigit()


def _excel_nummer(value):
    nummer = str(value).strip()
    if nummer.endswith('.0') and nummer[:-2].isdigit():
        return nummer[:-2]
    return nummer


def _excel_anzahl(value):
    """Liest eine positive ganze Schüleranzahl aus einer Excel-Zelle."""
    if value is None or str(value).strip() == '':
        return None
    try:
        anzahl = float(value)
    except (TypeError, ValueError):
        return None
    if not anzahl.is_integer() or anzahl < 1:
        return None
    return int(anzahl)


def _ermittle_zeilen_ohne_kopfzeile(workbook):
    """Erkennt einfache Excel-Listen ohne Klasse-/Nummer-Überschriften."""
    for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(values_only=True))

        # Format: Jede Zeile enthält direkt Klasse und Nummer, z. B. "5a | 12".
        inferred_rows = []
        for row_number, row in enumerate(values, start=1):
            non_empty = [(index, value) for index, value in enumerate(row) if value not in (None, '')]
            class_value = next(
                ((index, value) for index, value in non_empty
                 if not _ist_excel_nummer(value) and len(str(value).strip()) <= 20),
                None
            )
            number_value = next(
                ((index, value) for index, value in non_empty
                 if _ist_excel_nummer(value) and index != (class_value[0] if class_value else -1)),
                None
            )
            if class_value and number_value:
                inferred_rows.append((row_number, str(class_value[1]).strip(), _excel_nummer(number_value[1])))
        if inferred_rows:
            return inferred_rows

        # Format: Klassen stehen oben als Spalten, darunter stehen die Nummern.
        for header_index, header_row in enumerate(values[:20]):
            class_columns = [
                index for index, value in enumerate(header_row)
                if value not in (None, '') and not _ist_excel_nummer(value)
                and len(str(value).strip()) <= 20
            ]
            inferred_columns = []
            for row_number, row in enumerate(values[header_index + 1:], start=header_index + 2):
                for index in class_columns:
                    if index < len(row) and _ist_excel_nummer(row[index]):
                        inferred_columns.append((row_number, str(header_row[index]).strip(), _excel_nummer(row[index])))
            if inferred_columns:
                return inferred_columns
    return []


@admin_bp.route('/import_students', methods=['GET', 'POST'])
@login_required
def import_students():
    """Liest Klassen und Nummern aus einer Excel-Datei ein."""
    if request.method == 'POST':
        upload = request.files.get('file')
        if not upload or not (upload.filename or '').lower().endswith('.xlsx'):
            flash('Bitte eine Excel-Datei im Format .xlsx auswählen.', 'danger')
            return redirect(url_for('admin.import_students'))

        try:
            workbook = load_workbook(upload, read_only=True, data_only=True)
            detected = _finde_excel_spalten(workbook)
            if detected is None:
                source_rows = _ermittle_zeilen_ohne_kopfzeile(workbook)
                if not source_rows:
                    flash('Die Excel-Datei enthält keine erkennbare Kombination aus Klasse und Nummer.', 'danger')
                    return redirect(url_for('admin.import_students'))
            else:
                sheet, values, header_index, class_index, number_index = detected
                source_rows = []
                for row_number, row_values in enumerate(values[header_index + 1:], start=header_index + 2):
                    class_group = str(row_values[class_index]).strip() if class_index < len(row_values) and row_values[class_index] is not None else ''
                    nummer_value = row_values[number_index] if number_index < len(row_values) else None
                    source_rows.append((row_number, class_group, nummer_value))

            valid_rows = []
            errors = []
            for row_number, class_group, count_value in source_rows:
                count = _excel_anzahl(count_value)
                if not class_group or count is None:
                    errors.append(f'Zeile {row_number}: Klasse und positive Schüleranzahl fehlen.')
                    continue
                valid_rows.extend(
                    (row_number, class_group, str(number))
                    for number in range(1, count + 1)
                )

            inserted, skipped = db.import_students(valid_rows)
            message = f'{inserted} Personen eingefügt.'
            if skipped:
                message += f' {len(skipped)} bereits vorhandene Zeilen übersprungen.'
            if errors:
                message += f' {len(errors)} fehlerhafte Zeilen übersprungen.'
            flash(message, 'success' if inserted or not errors else 'warning')
            return render_template('admin_import_students.html', skipped=skipped, errors=errors)
        except Exception as error:
            flash(f'Excel-Datei konnte nicht verarbeitet werden: {error}', 'danger')
            return redirect(url_for('admin.import_students'))

    return render_template('admin_import_students.html')


@admin_bp.route('/edit_student/<int:student_id>', methods=['GET', 'POST'])
@login_required
def edit_student(student_id):
    """
    Schüler-Informationen bearbeiten.
    Validiert erforderliche Felder.
    """
    student = db.get_student_by_id(student_id)
    if not student:
        flash('Schüler nicht gefunden.', 'danger')
        return redirect(url_for('admin.list_students'))
    
    if request.method == 'POST':
        class_group = request.form.get('class_group', '').strip()
        nummer = request.form.get('nummer', '').strip()
        ill = bool(request.form.get('ill'))
        
        if not all([class_group, nummer]):
            flash('Klasse und Nummer müssen angegeben werden.', 'danger')
            return redirect(url_for('admin.edit_student', student_id=student_id))
        
        try:
            # Aktualisiere nur die angegebenen Felder
            db.update_student(student_id, class_group=class_group, nummer=nummer, ill=ill)
            flash('Schüler erfolgreich aktualisiert.', 'success')
            return redirect(url_for('admin.list_students'))
        except Exception as e:
            flash(f'Fehler beim Aktualisieren: {e}', 'danger')
    
    return render_template('admin_edit_student.html', student=student)


@admin_bp.route('/delete_student/<int:student_id>', methods=['POST'])
@login_required
def delete_student(student_id):
    """Schüler löschen."""
    try:
        db.delete_student(student_id)
        flash('Schüler erfolgreich gelöscht.', 'success')
    except Exception as e:
        flash(f'Fehler beim Löschen: {e}', 'danger')
    return redirect(url_for('admin.list_students'))


# ============================================================================
# LAUF-VERANSTALTUNGEN
# ============================================================================

@admin_bp.route('/laufgruppen', methods=['GET', 'POST'])
@login_required
def list_run_groups():
    """Laufgruppen mit mehreren Klassen anlegen und verwalten."""
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        class_groups = request.form.getlist('class_groups')
        try:
            db.create_run_group(name, class_groups)
            flash('Laufgruppe erfolgreich erstellt.', 'success')
            return redirect(url_for('admin.list_run_groups'))
        except Exception as error:
            flash(f'Laufgruppe konnte nicht erstellt werden: {error}', 'danger')
    return render_template(
        'admin_laufgruppen.html',
        run_groups=db.get_run_groups(),
        classes=db.get_classes()
    )


@admin_bp.route('/laufgruppen/<int:run_group_id>/delete', methods=['POST'])
@login_required
def delete_run_group(run_group_id):
    """Laufgruppe und zugehörige Messungen löschen."""
    try:
        with db.get_db_connection() as conn:
            conn.execute("DELETE FROM measurements WHERE run_group_id = ?", (run_group_id,))
            conn.execute("DELETE FROM run_groups WHERE id = ?", (run_group_id,))
            conn.commit()
        flash('Laufgruppe gelöscht.', 'success')
    except Exception as error:
        flash(f'Laufgruppe konnte nicht gelöscht werden: {error}', 'danger')
    return redirect(url_for('admin.list_run_groups'))

@admin_bp.route('/laeufe')
@login_required
def list_laeufe():
    """Alte Einzel-Laufseite auf die neue Laufgruppenverwaltung umleiten."""
    return redirect(url_for('admin.list_run_groups'))

@admin_bp.route('/erstellung_lauf', methods=['GET', 'POST'])
@login_required
def erstelle_lauf():
    """
    Neuen Lauf erstellen.
    Validiert erforderliche Felder.
    """
    return redirect(url_for('admin.list_run_groups'))


@admin_bp.route('/start_lauf', methods=['GET', 'POST'])
@login_required
def start_lauf():
    """
    Neuen Lauf starten.
    Schlägt verfügbare Klassen/Gruppen vor.
    """
    return redirect(url_for('admin.list_run_groups'))


@admin_bp.route('/end_lauf', methods=['POST'])
@login_required
def end_lauf():
    """Lauf beenden."""
    lauf_id = request.form.get('lauf_id')
    if not lauf_id:
        flash('Lauf-ID erforderlich.', 'danger')
        return redirect(url_for('admin.list_laeufe'))
    
    try:
        db.end_lauf(int(lauf_id))
        flash('Lauf beendet.', 'success')
    except Exception as e:
        flash(f'Fehler: {e}', 'danger')
    return redirect(url_for('admin.list_laeufe'))

#Restart Lauf
@admin_bp.route('/restart_lauf', methods=['POST'])
@login_required
def restart_lauf():
    """Lauf neustarten."""
    lauf_id = request.form.get('lauf_id')
    if not lauf_id:
        flash('Lauf-ID erforderlich.', 'danger')
        return redirect(url_for('admin.list_laeufe'))
    
    try:
        db.restart_lauf(int(lauf_id))
        flash('Lauf neu gestartet.', 'success')
    except Exception as e:
        flash(f'Fehler: {e}', 'danger')
    return redirect(url_for('admin.list_laeufe'))

@admin_bp.route('/delete_lauf', methods=['POST'])
@login_required
def delete_lauf():
    """Lauf löschen."""
    lauf_id = request.form.get('lauf_id')
    if not lauf_id:
        flash('Lauf-ID erforderlich.', 'danger')
        return redirect(url_for('admin.list_laeufe'))
    
    try:
        db.delete_lauf(int(lauf_id))
        flash('Lauf gelöscht.', 'success')
    except Exception as e:
        flash(f'Fehler: {e}', 'danger')
    return redirect(url_for('admin.list_laeufe'))
# ============================================================================
# ZEITMESSUNGEN
# ============================================================================

@admin_bp.route('/measurements')
@login_required
def list_measurements():
    """Zeitmessungs-Übersicht: Letzte 100 Messungen anzeigen."""
    measurements = db.get_measurements(limit=100)
    return render_template('admin_measurements.html', measurements=measurements)


@admin_bp.route('/clear_measurements', methods=['POST'])
@login_required
def clear_measurements():
    """Alle Zeitmessungen löschen."""
    try:
        db.clear_measurements()
        flash('Alle Messungen wurden gelöscht.', 'success')
    except Exception as e:
        flash(f'Fehler beim Löschen der Messungen: {e}', 'danger')
    return redirect(url_for('admin.list_measurements'))


# ============================================================================
# API-ENDPUNKTE (JSON)
# ============================================================================

@admin_bp.route('/api/students')
@login_required
def api_students():
    """API: Alle Schüler als JSON (inkl. kranke)."""
    return jsonify(db.get_students(include_ill=True))


@admin_bp.route('/api/measurements')
@login_required
def api_measurements():
    """API: Alle Zeitmessungen als JSON."""
    return jsonify(db.get_measurements())


@admin_bp.route('/api/laufgruppen')
@login_required
def api_run_groups():
    """API: Laufgruppen für das Admin-Interface."""
    return jsonify(db.get_run_groups())